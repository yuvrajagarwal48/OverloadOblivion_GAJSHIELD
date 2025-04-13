import os
import sys
import hashlib
import magic
import pandas as pd
import zipfile
import re
import logging
import json
from datetime import datetime
from openpyxl import load_workbook
import olefile
from pathlib import Path
import base64
import struct
import binascii
import math

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler("malware_detection.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class MalwareDetector:
    def __init__(self, verbose=False):
        self.verbose = verbose
        self.signatures = {
            "known_malware_md5": [
                # Add known malicious file hashes here
                "e6ff1bf0821f00d55bef70c5b12b3164",  # Example malicious Excel hash
                "a3c31d62dbaa48337c2ffd5e1093bcd7",  # Example malicious CSV hash
                # Add more known malware hashes here
            ]
        }
        
        self.suspicious_patterns = {
            'office_exploit_patterns': [
                # Excel 4.0 macros (XLM)
                re.compile(r'=EXEC\(', re.IGNORECASE),
                re.compile(r'=SYSTEM\(', re.IGNORECASE),
                re.compile(r'=SHELL\(', re.IGNORECASE),
                re.compile(r'=CMD\(', re.IGNORECASE),
                re.compile(r'=CALL\(', re.IGNORECASE),
                re.compile(r'=FORMULA\(', re.IGNORECASE),
                re.compile(r'=RUN\(', re.IGNORECASE),
                re.compile(r'=REGISTER\(', re.IGNORECASE),
                re.compile(r'=DDE\(', re.IGNORECASE),
                re.compile(r'=DDEAUTO\(', re.IGNORECASE),
                re.compile(r'=AUTO_OPEN\(', re.IGNORECASE),
                re.compile(r'=Workbook_Open\(', re.IGNORECASE),
                re.compile(r'powershell\s+', re.IGNORECASE),
                re.compile(r'cmd\.exe', re.IGNORECASE),
            ],
            
            # CVE patterns for common exploits
            'cve_patterns': [
                re.compile(r'CVE-2017-11882', re.IGNORECASE),  # Equation Editor
                re.compile(r'CVE-2018-0802', re.IGNORECASE),   # Equation Editor
                re.compile(r'CVE-2017-8570', re.IGNORECASE),   # Composite Moniker
                re.compile(r'CVE-2017-0199', re.IGNORECASE),   # HTA handler
                re.compile(r'CVE-2012-0158', re.IGNORECASE),   # MSCOMCTL
                re.compile(r'CVE-2015-1641', re.IGNORECASE),   # RTF
            ],
            
            # URL and network IOCs
            'network_ioc': [
                re.compile(r'https?://[^\s/$.?#].[^\s]*', re.IGNORECASE),  # URLs
                re.compile(r'(\\\\|//)[^\s/$.?#].[^\s]*', re.IGNORECASE),  # Network paths
                re.compile(r'mshta\.exe', re.IGNORECASE),                   # HTA handler
                re.compile(r'certutil\.exe', re.IGNORECASE),                # Often used for downloads
                re.compile(r'bitsadmin', re.IGNORECASE),                    # Transfer utility
            ],
            
            # Suspicious file extensions within content
            'suspicious_extensions': [
                re.compile(r'\.(exe|dll|vbs|js|ps1|bat|cmd|hta|msi|scr)\b', re.IGNORECASE),
            ],
            
            # Common obfuscation techniques
            'obfuscation': [
                re.compile(r'chr\(\d+\)', re.IGNORECASE),          # Character code
                re.compile(r'base64', re.IGNORECASE),              # Base64 encoding
                re.compile(r'cariage', re.IGNORECASE),             # Common misspelling indicating evasion
                re.compile(r'char\s*\(\s*\d+\s*\)', re.IGNORECASE), # VBA char function
                re.compile(r'\\u00[0-9a-f]{2}', re.IGNORECASE),    # Unicode escape
                # Look for long hex or decimal strings (potential shellcode)
                re.compile(r'(0x[0-9a-f]{2},?\s*){10,}', re.IGNORECASE),
            ]
        }
        
        self.risk_score = 0
        self.findings = []
        self.detailed_analysis = {}
    
    def analyze_file(self, file_path):
        """Main analysis method to check a file for malicious traits"""
        print(f"\n{'='*70}")
        print(f"MALWARE ANALYSIS REPORT - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*70}")
        print(f"File: {os.path.abspath(file_path)}")
        
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            print(f"ERROR: File not found: {file_path}")
            return False, "File not found"
        
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            self.add_finding("Warning", "File is empty (0 bytes)")
            return self.get_result()
            
        # Reset for new file
        self.risk_score = 0
        self.findings = []
        self.detailed_analysis = {
            "file_info": {},
            "content_analysis": {},
            "structure_analysis": {},
            "suspicious_elements": [],
            "summary": {}
        }
        
        # Start timer for performance measurement
        start_time = datetime.now()
        
        try:
            # File metadata collection
            file_info = self._collect_file_metadata(file_path)
            self.detailed_analysis["file_info"] = file_info
            
            # Print basic file info
            print(f"\nFILE INFORMATION:")
            print(f"  Name: {file_info['filename']}")
            print(f"  Size: {file_info['size_readable']} ({file_info['size']} bytes)")
            print(f"  Type: {file_info['detected_type']}")
            print(f"  MD5:  {file_info['md5']}")
            print(f"  SHA256: {file_info['sha256']}")
            print(f"  Created: {file_info['created_time']}")
            print(f"  Modified: {file_info['modified_time']}")
            
            # Basic file checks
            self._check_file_extension(file_path)
            self._check_file_signature(file_path)
            self._check_known_hash(file_info['md5'], file_info['sha256'])
            
            # Format-specific checks
            file_type = file_info['detected_type'].lower()
            
            if "excel" in file_type or file_path.lower().endswith(".xlsx"):
                self._analyze_xlsx(file_path)
            elif "csv" in file_type or file_path.lower().endswith(".csv"):
                self._analyze_csv(file_path)
            else:
                self._analyze_unknown(file_path)
                
            # Generic pattern checks
            self._check_patterns(file_path)
            
            # Collect entropy statistics
            self._analyze_entropy(file_path)
            
            # Analyze embedded objects
            self._check_for_embedded_objects(file_path)
            
            # End timer
            end_time = datetime.now()
            analysis_time = (end_time - start_time).total_seconds()
            
            # Build summary
            result = self.get_result()
            result["analysis_time"] = analysis_time
            
            # Print detailed findings
            self._print_detailed_report(result)
            
            # Save report to file
            self._save_report(file_path, result)
            
            return result
            
        except Exception as e:
            logger.error(f"Error analyzing file: {str(e)}", exc_info=True)
            self.add_finding("Error", f"Analysis error: {str(e)}")
            result = self.get_result()
            self._print_detailed_report(result)
            return result
    
    def _collect_file_metadata(self, file_path):
        """Collect basic file metadata"""
        file_info = {}
        
        # Basic file properties
        file_info['path'] = os.path.abspath(file_path)
        file_info['filename'] = os.path.basename(file_path)
        file_info['extension'] = os.path.splitext(file_path)[1].lower()
        file_info['size'] = os.path.getsize(file_path)
        file_info['size_readable'] = self._format_size(file_info['size'])
        
        # File timestamps
        file_info['created_time'] = datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
        file_info['modified_time'] = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
        
        # File hashes
        with open(file_path, 'rb') as f:
            content = f.read()
            file_info['md5'] = hashlib.md5(content).hexdigest()
            file_info['sha256'] = hashlib.sha256(content).hexdigest()
        
        # File magic type
        try:
            file_info['magic_type'] = magic.from_file(file_path)
            file_info['magic_mime'] = magic.from_file(file_path, mime=True)
        except:
            file_info['magic_type'] = "Unknown"
            file_info['magic_mime'] = "Unknown"
        
        # Determine type based on extension and magic
        if "excel" in file_info['magic_type'].lower() or file_info['extension'] == ".xlsx":
            file_info['detected_type'] = "Microsoft Excel Document"
        elif "csv" in file_info['magic_type'].lower() or file_info['extension'] == ".csv":
            file_info['detected_type'] = "Comma-Separated Values (CSV)"
        elif zipfile.is_zipfile(file_path) and file_info['extension'] == ".xlsx":
            file_info['detected_type'] = "Microsoft Excel Document"
        else:
            file_info['detected_type'] = file_info['magic_type']
        
        return file_info
    
    def _format_size(self, size_bytes):
        """Format file size in human-readable format"""
        if size_bytes < 1024:
            return f"{size_bytes} bytes"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes/1024:.2f} KB"
        elif size_bytes < 1024 * 1024 * 1024:
            return f"{size_bytes/(1024*1024):.2f} MB"
        else:
            return f"{size_bytes/(1024*1024*1024):.2f} GB"
    
    def _check_file_extension(self, file_path):
        """Check if file extension matches content"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in ['.csv', '.xlsx']:
            self.add_finding("Suspicious", f"File has unexpected extension: {ext}")
            self.risk_score += 10

    def _check_file_signature(self, file_path):
        """Check file magic bytes"""
        with open(file_path, 'rb') as f:
            header = f.read(16)  # Read more bytes for better identification
        
        # XLSX signature check (PK zip header)
        if header.startswith(b'PK\x03\x04'):
            if not file_path.lower().endswith('.xlsx'):
                self.add_finding("Suspicious", f"File has ZIP/XLSX signature but incorrect extension: {os.path.splitext(file_path)[1]}")
                self.risk_score += 15
        
        # CSV signature check (no specific signature, but check for text vs. binary)
        elif file_path.lower().endswith('.csv'):
            is_binary = False
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    sample = f.read(1024)
                    if '\0' in sample:
                        is_binary = True
            except UnicodeDecodeError:
                is_binary = True
                
            if is_binary:
                self.add_finding("Suspicious", "CSV file appears to contain binary data")
                self.risk_score += 25
        
        # Check for OLE (old Excel format)
        if header.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):
            self.add_finding("Info", "File has OLE Compound Document format (possible legacy Excel format)")
            if not any(file_path.lower().endswith(ext) for ext in ['.xls', '.doc', '.ppt']):
                self.add_finding("Suspicious", "OLE file format with unexpected extension")
                self.risk_score += 20
    
    def _check_known_hash(self, md5, sha256):
        """Check file hash against known malicious hashes"""
        if md5 in self.signatures["known_malware_md5"]:
            self.add_finding("Critical", f"File matches known malware hash (MD5): {md5}")
            self.risk_score += 100
            return True
        
        # Add more hash checks here if needed
        return False

    def _analyze_xlsx(self, file_path):
        """Analyze XLSX file for malicious content"""
        try:
            # Check if it's a valid zip file (XLSX is a zip container)
            if not zipfile.is_zipfile(file_path):
                self.add_finding("Critical", "File has XLSX extension but is not a valid ZIP container")
                self.risk_score += 50
                return
            
            xlsx_details = {"structure": {}, "content_types": [], "sheets": [], "external_refs": [], "macros": []}
            
            with zipfile.ZipFile(file_path) as zfile:
                # Get list of all files in the XLSX
                filelist = zfile.namelist()
                xlsx_details["structure"]["file_count"] = len(filelist)
                
                # List of important files to check
                important_files = [
                    "[Content_Types].xml", 
                    "xl/workbook.xml",
                    "xl/sharedStrings.xml",
                    "xl/styles.xml"
                ]
                
                # Check for required files
                missing_files = [f for f in important_files if f not in filelist]
                if missing_files:
                    self.add_finding("Suspicious", f"XLSX missing important files: {', '.join(missing_files)}")
                    self.risk_score += 30
                
                # Get file structure
                xlsx_details["structure"]["files"] = filelist
                
                # Check for suspicious file extensions within the archive
                suspicious_extensions = ['.exe', '.dll', '.vbs', '.ps1', '.bat', '.cmd', '.js', '.hta', '.jar', '.wsf']
                suspicious_files = []
                for item in filelist:
                    ext = os.path.splitext(item)[1].lower()
                    if ext in suspicious_extensions:
                        suspicious_files.append(item)
                        self.add_finding("Critical", f"XLSX contains suspicious file: {item}")
                        self.risk_score += 80
                
                if suspicious_files:
                    xlsx_details["structure"]["suspicious_files"] = suspicious_files
                
                # Check for macro presence
                if "vbaProject.bin" in filelist:
                    self.add_finding("Warning", "XLSX contains VBA macro code (vbaProject.bin)")
                    self.risk_score += 40
                    xlsx_details["macros"].append({"name": "vbaProject.bin", "type": "VBA"})
                    
                    # Try to extract macro information
                    try:
                        temp_dir = "temp_extracted"
                        os.makedirs(temp_dir, exist_ok=True)
                        
                        zfile.extract("vbaProject.bin", path=temp_dir)
                        macro_path = os.path.join(temp_dir, "vbaProject.bin")
                        
                        if os.path.exists(macro_path):
                            try:
                                # Try to analyze the macro content
                                ole = olefile.OleFile(macro_path)
                                macro_streams = []
                                
                                # List all streams that contain macro code
                                for stream in ole.listdir():
                                    stream_path = '/'.join(stream)
                                    if stream_path.startswith('VBA/'):
                                        stream_name = stream[-1]
                                        try:
                                            stream_content = ole.openstream(stream).read().decode('utf-8', errors='ignore')
                                            suspicious_terms = ['AutoOpen', 'Document_Open', 'Auto_Open', 'Workbook_Open', 
                                                              'Shell', 'WScript', 'powershell', 'cmd.exe', 'ActiveXObject']
                                            
                                            # Check for suspicious terms in macro
                                            found_terms = []
                                            for term in suspicious_terms:
                                                if term.lower() in stream_content.lower():
                                                    found_terms.append(term)
                                            
                                            if found_terms:
                                                self.add_finding("Critical", 
                                                              f"Suspicious terms in macro '{stream_name}': {', '.join(found_terms)}")
                                                self.risk_score += 60
                                                
                                            macro_streams.append({"name": stream_name, 
                                                               "suspicious_terms": found_terms if found_terms else None})
                                            
                                        except Exception as e:
                                            logger.debug(f"Error reading macro stream {stream_name}: {str(e)}")
                                
                                xlsx_details["macros"].append({"streams": macro_streams})
                                ole.close()
                            except Exception as e:
                                logger.debug(f"Error analyzing macro file: {str(e)}")
                            finally:
                                # Clean up
                                try:
                                    os.remove(macro_path)
                                    os.rmdir(temp_dir)
                                except:
                                    pass
                    except Exception as e:
                        self.add_finding("Info", f"Could not analyze macro content: {str(e)}")
                
                # Check for external references
                try:
                    external_refs_found = False
                    for filename in filelist:
                        if filename.startswith("xl/externalLinks/"):
                            external_refs_found = True
                            try:
                                with zfile.open(filename) as f:
                                    content = f.read().decode('utf-8', errors='ignore')
                                    # Look for URLs or file paths in external links
                                    urls = re.findall(r'https?://[^\s<>"\']+', content)
                                    file_paths = re.findall(r'[A-Za-z]:\\[^<>:"/\\|?*]+', content)
                                    
                                    if urls or file_paths:
                                        refs = urls + file_paths
                                        self.add_finding("Suspicious", 
                                                      f"XLSX contains external references: {', '.join(refs[:3])}{'...' if len(refs) > 3 else ''}")
                                        self.risk_score += 40
                                        xlsx_details["external_refs"].extend(refs)
                            except Exception as e:
                                logger.debug(f"Error analyzing external links: {str(e)}")
                    
                    if external_refs_found:
                        self.add_finding("Suspicious", "XLSX contains external data links")
                        self.risk_score += 30
                except Exception as e:
                    logger.debug(f"Error checking external links: {str(e)}")
                
                # Extract and check sheet content
                sheet_files = [f for f in filelist if f.startswith("xl/worksheets/sheet")]
                sheet_data = []
                
                for sheet_file in sheet_files:
                    try:
                        with zfile.open(sheet_file) as f:
                            content = f.read().decode('utf-8', errors='ignore')
                            
                            # Look for formula cells
                            formulas = re.findall(r'<f[^>]*>(.*?)</f>', content)
                            if formulas:
                                sheet_data.append({
                                    "name": sheet_file,
                                    "formula_count": len(formulas),
                                    "formulas": formulas[:5]  # Store first 5 formulas
                                })
                                
                                # Check for suspicious formulas
                                suspicious_formula_terms = ['cmd', 'exec', 'system', 'powershell', 'http', 'https', 'file://']
                                for formula in formulas:
                                    if any(term in formula.lower() for term in suspicious_formula_terms):
                                        self.add_finding("Critical", f"Potentially malicious formula found: {formula[:100]}")
                                        self.risk_score += 60
                    except Exception as e:
                        logger.debug(f"Error analyzing sheet {sheet_file}: {str(e)}")
                
                xlsx_details["sheets"] = sheet_data
                
            # Additional check for DDE formulas using openpyxl
            try:
                wb = load_workbook(file_path, read_only=True, data_only=False)
                sheets_with_formulas = []
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    has_suspicious_formula = False
                    formulas_found = []
                    
                    # Check the first 200 rows x 20 columns (to avoid processing entire large sheets)
                    for row in sheet.iter_rows(max_row=200, max_col=20):
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                cell_value = cell.value.lower()
                                
                                # Check if it's a formula
                                if cell_value.startswith("="):
                                    formulas_found.append(f"{cell.coordinate}: {cell.value}")
                                    
                                    # Check for suspicious content
                                    if any(term in cell_value for term in ['cmd', 'powershell', 'http', 'dde', 'exec']):
                                        self.add_finding("Critical", 
                                                       f"Potentially malicious formula in cell {cell.coordinate}: {cell.value}")
                                        self.risk_score += 60
                                        has_suspicious_formula = True
                    
                    if formulas_found:
                        sheets_with_formulas.append({
                            "name": sheet_name,
                            "has_suspicious_formula": has_suspicious_formula,
                            "formulas_sampled": formulas_found[:5]  # First 5 formulas only
                        })
                
                # Add to details
                xlsx_details["formula_analysis"] = {
                    "sheets_with_formulas": len(sheets_with_formulas),
                    "details": sheets_with_formulas
                }
                
            except Exception as e:
                self.add_finding("Warning", f"Could not check cell formulas with openpyxl: {str(e)}")
            
            # Add details to analysis output
            self.detailed_analysis["structure_analysis"]["xlsx"] = xlsx_details
            
        except Exception as e:
            self.add_finding("Error", f"Error analyzing XLSX file: {str(e)}")
    
    def _analyze_csv(self, file_path):
        """Analyze CSV file for malicious content"""
        try:
            # Initialize CSV details
            csv_details = {"structure": {}, "content": {}, "formula_threats": []}
            suspicious_lines = []
            formula_count = 0
            
            # Check for formula injection patterns
            formula_markers = ["=cmd", "=exec", "=powershell", "=https://", "=http://", "=file://", "=dnl", "=msexec", "=cmd|"]
            
            # Count lines and analyze content
            line_count = 0
            field_counts = []
            
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for i, line in enumerate(f, 1):
                    line_count += 1
                    
                    # Count fields (naively split by comma)
                    fields = line.count(',') + 1
                    field_counts.append(fields)
                    
                    # Check for formula injection
                    if any(marker in line.lower() for marker in formula_markers):
                        suspicious_lines.append((i, line.strip()))
                        formula_count += 1
                        
                        # Limit collection to avoid memory issues with large files
                        if len(suspicious_lines) >= 10:
                            break
            
            csv_details["structure"]["line_count"] = line_count
            csv_details["structure"]["field_count_min"] = min(field_counts) if field_counts else 0
            csv_details["structure"]["field_count_max"] = max(field_counts) if field_counts else 0
            csv_details["structure"]["field_count_consistent"] = (min(field_counts) == max(field_counts)) if field_counts else False
            
            # Report on formula threats
            if suspicious_lines:
                csv_details["formula_threats"] = [{"line": line_num, "content": content[:100]} 
                                                for line_num, content in suspicious_lines]
                
                for line_num, content in suspicious_lines:
                    self.add_finding("Critical", f"Potential formula injection at line {line_num}: {content[:100]}...")
                    self.risk_score += 70
            
            # Check if this is actually a CSV by looking at structure
            try:
                # Use pandas with a small sample size to avoid loading large files
                df = pd.read_csv(file_path, nrows=10)
                csv_details["content"]["columns"] = list(df.columns)
                csv_details["content"]["rows_sampled"] = len(df)
                csv_details["content"]["detected_column_count"] = len(df.columns)
                
                if df.shape[1] <= 1:  # Only one column might indicate it's not really CSV
                    self.add_finding("Suspicious", "File has CSV extension but does not appear to be comma-separated")
                    self.risk_score += 15
            except Exception as e:
                self.add_finding("Suspicious", f"File has CSV extension but could not be parsed as CSV: {str(e)}")
                self.risk_score += 25
                csv_details["content"]["parse_error"] = str(e)
            
            # Check for embedded URIs/URLs
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read(10240)  # Read first 10KB
                    urls = re.findall(r'https?://[^\s,"\'\r\n]+', content)
                    
                    if urls:
                        csv_details["content"]["urls_detected"] = urls[:10]  # First 10 URLs
                        
                        # Check for suspicious domains or patterns
                        suspicious_domains = ['.xyz', '.top', '.ru', '.cn', 'pastebin', 'raw.githubusercontent']
                        suspicious_urls = [url for url in urls if any(domain in url.lower() for domain in suspicious_domains)]
                        
                        if suspicious_urls:
                            self.add_finding("Suspicious", f"CSV contains potentially suspicious URLs: {', '.join(suspicious_urls[:3])}")
                            self.risk_score += 30
            except Exception as e:
                logger.debug(f"Error checking for URLs in CSV: {str(e)}")
            
            # Add to detailed analysis
            self.detailed_analysis["structure_analysis"]["csv"] = csv_details
            
        except Exception as e:
            self.add_finding("Error", f"Error analyzing CSV file: {str(e)}")
    
    def _analyze_unknown(self, file_path):
        """Analyze unknown file types for malicious content"""
        try:
            # Try to determine if it's actually a CSV or XLSX with wrong extension
            with open(file_path, 'rb') as f:
                header = f.read(8)
            
            # Check for ZIP signature (potential XLSX)
            if header.startswith(b'PK\x03\x04'):
                try:
                    with zipfile.ZipFile(file_path) as zf:
                        if "[Content_Types].xml" in zf.namelist() and any(name.startswith("xl/") for name in zf.namelist()):
                            self.add_finding("Suspicious", "File appears to be XLSX but has incorrect extension")
                            self._analyze_xlsx(file_path)
                            return
                except:
                    pass
            
            # Check if it might be a CSV
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    first_lines = [f.readline() for _ in range(3)]
                    comma_count = sum(line.count(',') for line in first_lines)
                    if comma_count > 5:  # Arbitrary threshold
                        self.add_finding("Suspicious", "File appears to be CSV but has incorrect extension")
                        self._analyze_csv(file_path)
                        return
            except:
                pass
            
            self.add_finding("Warning", "File type is not recognized as CSV or XLSX")
            self.risk_score += 15
            
            # Generic binary analysis
            with open(file_path, 'rb') as f:
                content = f.read()
                
            # Look for suspicious strings
            text_content = content.decode('latin-1')  # Use latin-1 to avoid decode errors
            suspicious_strings = []
            
            for pattern_name, patterns in self.suspicious_patterns.items():
                for pattern in patterns:
                    matches = pattern.findall(text_content)
                    if matches:
                        for match in matches[:5]:  # Limit to 5 matches per pattern
                            suspicious_strings.append({
                                "pattern_type": pattern_name,
                                "pattern": pattern.pattern,
                                "match": match if len(str(match)) < 100 else str(match)[:100] + "..."
                            })
                            
                            self.add_finding("Suspicious", f"Suspicious pattern '{pattern_name}' found: {match[:50]}...")
                            self.risk_score += 25
            
            self.detailed_analysis["structure_analysis"]["unknown_file"] = {
                "suspicious_strings_found": len(suspicious_strings),
                "suspicious_strings": suspicious_strings[:20]  # Limit to 20 entries
            }
            
        except Exception as e:
            self.add_finding("Error", f"Error analyzing unknown file type: {str(e)}")
    
    def _check_patterns(self, file_path):
        """Check file against suspicious patterns"""
        try:
            # Read file content (binary mode)
            with open(file_path, 'rb') as f:
                content = f.read()
                
            # Convert to string for regex matching
            text_content = content.decode('latin-1')  # Use latin-1 to avoid decode errors
            
            # Track matches for reporting
            all_matches = []
            
            # Check for all pattern categories
            for category_name, patterns in self.suspicious_patterns.items():
                category_matches = []
                
                for pattern in patterns:
                    matches = pattern.findall(text_content)
                    
                    if matches:
                        sample_matches = matches[:3]  # Limit to first 3 matches
                        
                        for match in sample_matches:
                            match_str = match if isinstance(match, str) else str(match)
                            category_matches.append({
                                "pattern": pattern.pattern,
                                "match": match_str[:100] + "..." if len(match_str) > 100 else match_str
                            })
                
                if category_matches:
                    all_matches.append({
                        "category": category_name,
                        "match_count": len(category_matches),
                        "samples": category_matches
                    })
                    
                    # Add findings based on category
                    severity = "Critical" if category_name in ['office_exploit_patterns', 'cve_patterns'] else "Suspicious"
                    risk_points = 60 if severity == "Critical" else 30
                    
                    self.add_finding(severity, f"Found {len(category_matches)} matches for {category_name} pattern")
                    self.risk_score += risk_points
            
            # Save detailed matches to analysis output
            if all_matches:
                self.detailed_analysis["content_analysis"]["pattern_matches"] = all_matches
        
        except Exception as e:
            self.add_finding("Error", f"Error checking file patterns: {str(e)}")
    
    def _analyze_entropy(self, file_path):
        """Calculate entropy for file sections to detect obfuscation or encryption"""
        try:
            # Read the file content
            with open(file_path, 'rb') as f:
                content = f.read()
            
            if len(content) == 0:
                return
            
            # Divide file into sections for entropy analysis
            section_size = min(4096, len(content))
            num_sections = min(10, len(content) // section_size)
            
            if num_sections == 0:
                num_sections = 1
                section_size = len(content)
            
            entropy_values = []
            
            # Calculate entropy for each section
            for i in range(num_sections):
                start = i * section_size
                end = min(start + section_size, len(content))
                section = content[start:end]
                
                # Calculate entropy
                byte_counts = {}
                for byte in section:
                    byte_counts[byte] = byte_counts.get(byte, 0) + 1
                
                entropy = 0
                for count in byte_counts.values():
                    probability = count / len(section)
                    entropy -= probability * (math.log(probability) / math.log(2))
                
                entropy_values.append(entropy)
            
            # Analyze entropy values
            avg_entropy = sum(entropy_values) / len(entropy_values)
            max_entropy = max(entropy_values)
            
            # Add entropy data to detailed analysis
            self.detailed_analysis["content_analysis"]["entropy"] = {
                "average": avg_entropy,
                "maximum": max_entropy,
                "values": entropy_values,
                "high_entropy_sections": [i for i, e in enumerate(entropy_values) if e > 7.0]
            }
            
            # High entropy could indicate encryption, compression, or obfuscation
            if max_entropy > 7.5:
                self.add_finding("Suspicious", f"High entropy detected (max: {max_entropy:.2f}), possible encryption or obfuscation")
                self.risk_score += 25
            elif avg_entropy > 6.8:
                self.add_finding("Info", f"Moderately high entropy detected (avg: {avg_entropy:.2f})")
                self.risk_score += 10
        
        except Exception as e:
            logger.debug(f"Error calculating entropy: {str(e)}")
    
    def _check_for_embedded_objects(self, file_path):
        """Check for embedded OLE objects or other suspicious content"""
        try:
            # Check for OLE objects
            with open(file_path, 'rb') as f:
                content = f.read()
            
            embedded_objects = []
            
            # Look for OLE magic bytes
            ole_positions = [m.start() for m in re.finditer(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1', content)]
            if ole_positions:
                for pos in ole_positions[:5]:  # Limit to first 5 occurrences
                    embedded_objects.append({
                        "type": "OLE Object",
                        "offset": pos,
                        "header": binascii.hexlify(content[pos:pos+16]).decode('ascii')
                    })
                    
                    self.add_finding("Suspicious", f"Embedded OLE object found at offset {pos}")
                    self.risk_score += 40
            
            # Look for executable headers (MZ)
            exe_positions = [m.start() for m in re.finditer(b'MZ', content)]
            if exe_positions:
                for pos in exe_positions[:5]:
                    # Check for PE header structure
                    if pos + 64 < len(content):
                        try:
                            # Check for PE header at standard offset
                            pe_offset = struct.unpack("<I", content[pos+0x3C:pos+0x40])[0]
                            if pos + pe_offset + 2 < len(content) and content[pos+pe_offset:pos+pe_offset+2] == b'PE':
                                embedded_objects.append({
                                    "type": "PE Executable",
                                    "offset": pos,
                                    "pe_header_offset": pe_offset,
                                    "header": binascii.hexlify(content[pos:pos+16]).decode('ascii')
                                })
                                
                                self.add_finding("Critical", f"Embedded executable (PE) found at offset {pos}")
                                self.risk_score += 90
                        except:
                            pass
            
            # Check for Base64-encoded content that might be executable
            base64_pattern = re.compile(b'(?:[A-Za-z0-9+/]{4}){8,}(?:[A-Za-z0-9+/]{2}==|[A-Za-z0-9+/]{3}=)?', re.MULTILINE)
            base64_chunks = base64_pattern.findall(content)
            
            if base64_chunks:
                chunks_to_check = base64_chunks[:3]  # Check up to 3 potential base64 chunks
                
                for i, chunk in enumerate(chunks_to_check):
                    if len(chunk) > 100:  # Only check longer chunks
                        try:
                            # Try to decode and check for executable or script signatures
                            decoded = base64.b64decode(chunk)
                            
                            # Check for common file signatures in the decoded content
                            if decoded.startswith(b'MZ') or decoded.startswith(b'PK\x03\x04') or b'<script' in decoded:
                                embedded_objects.append({
                                    "type": "Base64-encoded object",
                                    "length": len(chunk),
                                    "decoded_signature": binascii.hexlify(decoded[:16]).decode('ascii')
                                })
                                
                                self.add_finding("Critical", f"Base64-encoded object found ({len(decoded)} bytes when decoded)")
                                self.risk_score += 70
                        except:
                            pass
            
            if embedded_objects:
                self.detailed_analysis["suspicious_elements"].extend(embedded_objects)
        
        except Exception as e:
            logger.debug(f"Error checking for embedded objects: {str(e)}")
    
    def add_finding(self, severity, message):
        """Add a finding to the results"""
        self.findings.append({
            "severity": severity,
            "message": message,
            "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    
    def get_result(self):
        """Get final analysis result"""
        risk_level = "Low"
        if self.risk_score >= 80:
            risk_level = "Critical"
        elif self.risk_score >= 50:
            risk_level = "High"
        elif self.risk_score >= 20:
            risk_level = "Medium"
            
        return {
            "risk_score": self.risk_score,
            "risk_level": risk_level,
            "findings": self.findings,
            "detailed_analysis": self.detailed_analysis
        }
    
    def _print_detailed_report(self, result):
        """Display detailed report to console"""
        # Print horizontal rule
        print(f"\n{'-'*70}")
        print("ANALYSIS RESULTS")
        print(f"{'-'*70}")
        
        # Print risk level with alert formatting
        if result["risk_level"] in ["High", "Critical"]:
            print(f"\n⚠️  RISK LEVEL: {result['risk_level'].upper()} (Score: {result['risk_score']})")
        else:
            print(f"\nRISK LEVEL: {result['risk_level']} (Score: {result['risk_score']})")
        
        # Print findings grouped by severity
        print("\nFINDINGS:")
        severities = ["Critical", "Suspicious", "Warning", "Info", "Error"]
        
        for severity in severities:
            severity_findings = [f for f in result["findings"] if f["severity"] == severity]
            if severity_findings:
                print(f"\n  {severity} ({len(severity_findings)}):")
                for finding in severity_findings:
                    print(f"    - {finding['message']}")
        
        # Print analysis time
        if "analysis_time" in result:
            print(f"\nAnalysis completed in {result['analysis_time']:.2f} seconds")
        
        print(f"\n{'-'*70}")
    
    def _save_report(self, file_path, result):
        """Save detailed report to a file"""
        try:
            # Create report directory
            report_dir = "malware_scan_reports"
            os.makedirs(report_dir, exist_ok=True)
            
            # Create report filename based on input file
            base_name = os.path.basename(file_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_file = os.path.join(report_dir, f"report_{base_name}_{timestamp}.json")
            
            # Write report as JSON
            with open(report_file, 'w') as f:
                json.dump(result, f, indent=2, default=str)
            
            print(f"Detailed report saved to: {report_file}")
            
            # Also create a text summary
            summary_file = os.path.join(report_dir, f"summary_{base_name}_{timestamp}.txt")
            with open(summary_file, 'w') as f:
                f.write(f"Malware Analysis Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"File: {os.path.abspath(file_path)}\n")
                f.write(f"Risk Level: {result['risk_level']} (Score: {result['risk_score']})\n\n")
                
                f.write("FINDINGS:\n")
                severities = ["Critical", "Suspicious", "Warning", "Info", "Error"]
                
                for severity in severities:
                    severity_findings = [finding for finding in result["findings"] if finding["severity"] == severity]
                    if severity_findings:
                        f.write(f"\n{severity} ({len(severity_findings)}):\n")
                        for finding in severity_findings:
                            f.write(f"  - {finding['message']}\n")
            
            print(f"Summary report saved to: {summary_file}")
            
        except Exception as e:
            logger.error(f"Error saving report: {str(e)}")
            print(f"Error saving report: {str(e)}")


def main():
    # Import math here to avoid potential import errors at the top level
    import math
    
    print("=" * 70)
    print("CSV and XLSX Malware Detector")
    print("=" * 70)
    
    # Ask for file to scan
    print("\nPlease enter the path to the file you want to analyze:")
    file_path = input("> ").strip()
    
    if not file_path:
        print("No file specified. Exiting.")
        return 1
        
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' does not exist.")
        return 1
        
    if os.path.isdir(file_path):
        print(f"Error: '{file_path}' is a directory, not a file.")
        return 1
    
    # Create and run the detector
    detector = MalwareDetector(verbose=True)
    detector.analyze_file(file_path)
    
    print("\nPress Enter to exit...")
    input()
    
    return 0

# # Example Usage

# file_path=''
# detector = MalwareDetector(verbose=True)
# result=detector.analyze_file(file_path)

# print(result)